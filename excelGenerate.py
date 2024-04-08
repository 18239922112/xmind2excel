# -*- coding: utf-8 -*-
"""
@Time ： 2024/4/3 10:10
@Auth ： 七月
@File ：excelGenerate.py
@IDE ：PyCharm
"""
from openpyxl import Workbook,load_workbook
from makeLog import Logger
import os
log = Logger()

def excelGen(file_path):
    '''
    excel文件生成
    :return:
    '''
    workbook = Workbook()

    # 创建新的工作表
    sheet1 = workbook.active
    #设置表格标题
    sheet1.title = '测试用例'
    #定义表头
    sheet1['A1'] = '用例名称'
    sheet1['B1'] = '需求号'
    sheet1['C1'] = '测试范围'
    sheet1['D1'] = '模块名称'
    sheet1['E1'] = '优先级'
    sheet1['F1'] = '为核心用例'
    sheet1['G1'] = '已自动化'
    sheet1['H1'] = '为回归用例'
    sheet1['I1'] = '前提条件'
    sheet1['J1'] = '操作步骤'
    sheet1['K1'] = '期望结果'
    sheet1['L1'] = '执行结果'

    #保存至文件
    workbook.save(file_path)
    os.chmod(file_path,0o777)
    log.info('生成excel文件成功:'+file_path)
    #关闭链接
    workbook.close()

def writeExcel(file_path,parms):
    '''
    将xmind数据写入excel
    :return:
    '''
    workbook = load_workbook(file_path)

    sheet1 = workbook.active


    sheet1.append(parms)
    # 保存至文件
    workbook.save(file_path)
    # 关闭链接
    workbook.close()





























